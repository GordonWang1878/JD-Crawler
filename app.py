#!/usr/bin/env python3
"""
JD价格爬虫 - Web前端
"""
import os
import re
import time
import random
import glob as glob_mod
from io import BytesIO
from datetime import datetime
from threading import Thread
from flask import Flask, render_template, request, jsonify, send_file
from flask_socketio import SocketIO
from flask_cors import CORS
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from werkzeug.utils import secure_filename
# Patchright 版 JD crawler — 替代之前的 selenium+CDP attach 方案
# 2026-05 京东升级反爬,selenium 即便 CDP attach 也被秒拒,patchright 修补了底层指纹
from jd_crawler_patchright import JDCrawlerViaSearch, _is_chrome_running_on_cdp_port, CDP_PORT
from tmall_crawler import TmallCrawler, parse_tmall_item_id
import profile_provision

# 初始化Flask应用
app = Flask(__name__)
app.config['SECRET_KEY'] = 'jd-crawler-simple-2025'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['TEMPLATES_AUTO_RELOAD'] = True  # 改 templates/*.html 不用重启 Flask

# 确保目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

# 初始化扩展
CORS(app)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# ===== 限流参数(自动分批 + 批次间冷却,降低反爬触发率) =====
# 京东 PC 新版风控约在单会话 40-50 条触发,降到 30 条以内单批
JD_BATCH_SIZE = 25
JD_BATCH_COOLDOWN = 600    # 10 分钟

# 爬取强度预设 — UI 上让用户在「稳」和「快」之间权衡。冷却都是 10 分钟,
# 只调每批条数:批越大→批数越少→省下的全是 10 分钟的批间等待,但单会话连续请求越多、越易触发风控。
# ⚠️ 实测:enhanced(100/批)在第 3 批(~第 209 条)即触发京东 PC 频控页
#    (pc-frequent-pro.pf.jd.com/?reason=403),之后连续失败 100+ 条、完全进不去商品页 —— 不可用,已移除。
#    未知 speed 一律回退 regular。
JD_SPEED_PRESETS = {
    'regular':  {'batch_size': 25,  'cooldown': 600, 'label': '常规'},
    'fast':     {'batch_size': 50,  'cooldown': 600, 'label': '快速'},
}
TMALL_BATCH_SIZE = 25      # 天猫反爬同样严格
TMALL_BATCH_COOLDOWN = 1500  # 25 分钟

# ===== 全局状态 =====
# 串行锁:同一时刻只允许一个平台跑(避免两个 Chrome 资源冲突)
is_crawling = False
current_platform = None  # 'jd' or 'tmall' or None

# 共享展示数据
live_results = []  # 每条 row 含 'platform' 字段
current_results = []

# 京东专属
crawler_instance = None  # JD crawler
current_batch_file = None  # JD 当前输出文件
uploaded_df = None  # JD 上传的 dataframe(预览用)
uploaded_urls = []
uploaded_rows = []  # JD 解析后的行

# 天猫专属
tmall_crawler_instance = None
current_tmall_batch_file = None
uploaded_tmall_rows = []  # 天猫解析后的行

# ==================== 辅助函数 ====================

def _norm_col(name):
    """归一化列名用于匹配"""
    return str(name).strip().lower().replace(' ', '').replace('_', '')

def _parse_excel(filepath):
    """读取 Excel 并提取 5 个核心列,返回 row dict 列表"""
    df = pd.read_excel(filepath)
    col_map = {_norm_col(c): c for c in df.columns}

    def pick(*candidates):
        # 精确匹配
        for c in candidates:
            if _norm_col(c) in col_map:
                return col_map[_norm_col(c)]
        # 前缀匹配 fallback(兼容 "Price Reference_0" 之类的后缀列名)
        for c in candidates:
            nc = _norm_col(c)
            for norm, orig in col_map.items():
                if norm.startswith(nc):
                    return orig
        return None

    brand_col = pick('Brand', '品牌')
    item_col = pick('Item', '型号', 'Model')
    url_col = pick('URL', 'ProductUrl std', 'ProductUrl', '链接')
    key_col = pick('Product Key', 'ProductKey', 'SKU')
    ref_col = pick('Price Reference', 'PriceReference', '参考价')

    rows = []
    for _, r in df.iterrows():
        def val(col):
            if not col:
                return ''
            v = r[col]
            if pd.isna(v):
                return ''
            if isinstance(v, float):
                if v.is_integer():
                    return str(int(v))
                return f'{v:.2f}'
            return str(v).strip()

        brand = val(brand_col)
        item = val(item_col)
        url = val(url_col)
        key = val(key_col)
        ref = val(ref_col)

        # URL 缺失时,从 Product Key 构造
        if not url and key:
            url = f"https://item.jd.com/{key}.html"

        # 从 URL 提取 product_id(爬取用)
        product_id = ''
        if url:
            m = re.search(r'/(\d+)\.html', url)
            if m:
                product_id = m.group(1)
        if not product_id and key:
            product_id = key

        if not url and not product_id:
            continue  # 跳过无效行

        rows.append({
            'brand': brand,
            'item': item,
            'url': url,
            'product_key': key,
            'price_reference': ref,
            'product_id': product_id,
        })
    return rows


def _parse_tmall_excel(filepath):
    """读取天猫 Excel,返回 row dict 列表.

    爬取依据 = URL(必填),由 parse_tmall_item_id 从 URL 提取真实 tmall id.
    Excel 的 ProductKey 列 = 业务侧编号(选填,只用于显示/对照,不参与爬取).
    """
    df = pd.read_excel(filepath)
    col_map = {_norm_col(c): c for c in df.columns}

    def pick(*candidates):
        for c in candidates:
            if _norm_col(c) in col_map:
                return col_map[_norm_col(c)]
        for c in candidates:
            nc = _norm_col(c)
            for norm, orig in col_map.items():
                if norm.startswith(nc):
                    return orig
        return None

    brand_col = pick('BRAND', 'Brand', '品牌')
    item_col = pick('Item', '型号', 'Model')
    url_col = pick('ProductUrl tmall', 'ProductUrl', 'URL', '链接')
    key_col = pick('Product Key', 'ProductKey', 'item_id', 'SKU')  # 业务侧编号(选填)
    ref_col = pick('Price Reference', 'PriceReference', '参考价')

    rows = []
    for _, r in df.iterrows():
        def val(col):
            if not col:
                return ''
            v = r[col]
            if pd.isna(v):
                return ''
            if isinstance(v, float):
                if v.is_integer():
                    return str(int(v))
                return f'{v:.2f}'
            return str(v).strip()

        brand = val(brand_col)
        item = val(item_col)
        url = val(url_col)
        product_key = val(key_col)  # 业务侧编号
        ref = val(ref_col)

        if not url:
            continue  # URL 必填,无 URL 无法爬取

        # 从 URL 提取真实 tmall/taobao item id(爬取用)
        real_item_id = parse_tmall_item_id(url) or ''
        if not real_item_id:
            continue  # URL 不规范,无法提取 id

        rows.append({
            'brand': brand,
            'item': item,
            'url': url,
            'item_id': real_item_id,    # 真实 tmall id,爬取用
            'product_key': product_key,  # 业务侧编号,只用于显示
            'price_reference': ref,
        })
    return rows


def _batch_cooldown(seconds: int, platform: str) -> bool:
    """批次间冷却,可被用户停止打断.每分钟打一条日志(分钟数变化时).
    返回 True 表示正常完成,False 表示被用户中止.
    """
    global is_crawling
    import math
    end_time = time.time() + seconds
    last_logged_min = None
    while time.time() < end_time:
        if not is_crawling:
            return False
        remaining = int(end_time - time.time())
        mins = remaining // 60
        secs = remaining % 60
        # 进度条:5 秒粒度更新(精确倒计时)
        emit_progress({
            'current_url': f'⏳ 批次间冷却中,剩余 {mins} 分 {secs} 秒(降低反爬触发率)',
            'platform': platform,
        })
        # 日志:仅在"显示分钟数"变化时打一条(向上取整,体感更准确)
        display_min = max(1, math.ceil(remaining / 60))
        if last_logged_min != display_min:
            emit_log('INFO', f'⏳ 批次间冷却中,剩余 {display_min} 分钟...', platform=platform)
            last_logged_min = display_min
        # 短 sleep 让 is_crawling 中止能快速响应
        time.sleep(min(5, max(1, remaining)))
    return is_crawling


def emit_log(level, message, platform=None):
    """发送日志到前端,可选平台前缀"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    prefix = ''
    if platform == 'jd':
        prefix = '[JD] '
    elif platform == 'tmall':
        prefix = '[Tmall] '
    full_message = f'{prefix}{message}'
    socketio.emit('log', {
        'timestamp': timestamp,
        'level': level,
        'message': full_message,
        'platform': platform,
    })
    print(f"[{timestamp}] [{level}] {full_message}")

def emit_progress(data):
    """发送进度更新到前端"""
    socketio.emit('progress', data)

def emit_result_row(row):
    """发送单条爬取结果到前端"""
    socketio.emit('result_row', row)

# ==================== 路由 ====================

@app.route('/')
def index():
    """首页"""
    return render_template('batch.html')

@app.route('/api/upload', methods=['POST'])
def api_upload():
    """上传Excel文件"""
    global uploaded_df, uploaded_urls, uploaded_rows

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Invalid file type'}), 400

    # 保存文件
    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    # 读取并解析
    try:
        df = pd.read_excel(filepath)
        uploaded_df = df

        rows = _parse_excel(filepath)
        uploaded_rows = rows
        uploaded_urls = [r['url'] for r in rows]

        return jsonify({
            'success': True,
            'filename': filename,
            'filepath': filepath,
            'url_count': len(rows),
            'columns': list(df.columns)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/preview')
def api_preview():
    """预览上传的数据 — 返回 Brand / Item / URL / Product Key / Price Reference"""
    global uploaded_rows

    if not uploaded_rows:
        return jsonify({'error': 'No file uploaded'}), 400

    rows = []
    for i, r in enumerate(uploaded_rows):
        rows.append({
            'index': i + 1,
            'brand': r.get('brand', ''),
            'item': r.get('item', ''),
            'url': r.get('url', ''),
            'product_key': r.get('product_key', ''),
            'price_reference': r.get('price_reference', ''),
        })

    return jsonify({'success': True, 'rows': rows, 'total': len(rows)})

@app.route('/api/history')
def api_history():
    """获取历史爬取结果文件列表"""
    output_dir = app.config['OUTPUT_FOLDER']
    files = []

    for f in sorted(glob_mod.glob(os.path.join(output_dir, '*.xlsx')), key=os.path.getmtime, reverse=True):
        fname = os.path.basename(f)
        stat = os.stat(f)
        # Try to read row count
        try:
            df = pd.read_excel(f)
            row_count = len(df)
        except:
            row_count = None

        files.append({
            'filename': fname,
            'size': stat.st_size,
            'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M'),
            'row_count': row_count
        })

    return jsonify({'success': True, 'files': files[:20]})  # Limit to 20 most recent

@app.route('/api/results')
def api_results():
    """获取当前爬取结果"""
    return jsonify({'success': True, 'results': live_results})

@app.route('/api/crawl/start', methods=['POST'])
def api_crawl_start():
    """开始批量爬取(京东)"""
    global is_crawling, current_platform, current_batch_file, live_results

    if is_crawling:
        return jsonify({'error': f'Crawler is already running ({current_platform})'}), 400

    if profile_provision.is_busy():
        return jsonify({'error': '正在配置账号(扫码/验证),请完成后再开始爬取'}), 400

    # 预检:profile 池必须非空(patchright 用 launch_persistent_context 直接接管 profile)
    import jd_profile_pool
    profiles = jd_profile_pool.list_available_profiles()
    if not profiles:
        return jsonify({
            'error': ('JD profile 池为空.请先在项目目录终端运行(一次):\n'
                      '    python3 prepare_jd_profile_pool_patchright.py 3\n\n'
                      '按提示给每个 profile 扫码登录京东.')
        }), 400

    data = request.json
    filepath = data.get('filepath')
    config = data.get('config', {})

    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'Invalid file path'}), 400

    # 生成输出文件名(JD_ 前缀,便于历史记录按平台识别)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"JD_Price_Marks_{timestamp}.xlsx"
    current_batch_file = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)

    # Reset live results — 只保留另一平台的(京东开跑时清掉旧京东结果)
    live_results = [r for r in live_results if r.get('platform') != 'jd']

    # 启动爬取任务
    is_crawling = True
    current_platform = 'jd'
    crawling_task = Thread(target=run_crawl_task, args=(filepath, current_batch_file, config))
    crawling_task.start()

    return jsonify({
        'success': True,
        'message': 'Crawling started',
        'output_file': output_filename
    })

@app.route('/api/crawl/retry', methods=['POST'])
def api_crawl_retry():
    """重试失败的商品(京东) — 覆盖原 Excel,不生成新文件,
    避免用户在历史记录里看到 _retry_ 和原文件两个版本搞混"""
    global is_crawling, current_platform, current_batch_file, live_results

    if is_crawling:
        return jsonify({'error': f'Crawler is already running ({current_platform})'}), 400

    # 收集可重试的京东项(含 skipped — 批次内主动跳过的)
    # 不在这里提前从 live_results 删除它们 — 旧逻辑在开爬「前」就删,一旦爬取线程崩溃,
    # 失败项既没重爬又已丢失、再也重试不了。改为保留,run_crawl_task_from_rows 用
    # _upsert_live_result 在拿到新结果时按 URL 替换旧记录(幂等),不重复也不丢。
    failed_items = [r for r in live_results
                    if r.get('platform') == 'jd'
                    and r.get('status') in RETRYABLE_STATUSES]

    if failed_items:
        # in-session:复用当前主文件名,retry 完成后回填得到完整最终版
        if not current_batch_file or not os.path.exists(os.path.dirname(current_batch_file) or '.'):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            current_batch_file = os.path.join(app.config['OUTPUT_FOLDER'], f"JD_Price_Marks_{timestamp}.xlsx")
    else:
        # 内存里没有(Flask 重启后)→ 从磁盘最新的错误小文件恢复:
        # 读它拿到「要重爬哪些」,主文件路径由文件名配对推出(去掉 _errors 后缀)。
        err_file = _latest_errors_file()
        if not err_file:
            return jsonify({'error': 'No failed JD items to retry'}), 400
        failed_items = _parse_excel(err_file)
        if not failed_items:
            return jsonify({'error': 'No failed JD items to retry'}), 400
        master = err_file[:-len('_errors.xlsx')] + '.xlsx'
        current_batch_file = master if os.path.exists(master) else os.path.join(
            app.config['OUTPUT_FOLDER'],
            f"JD_Price_Marks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        emit_log('INFO', f'内存无失败项,从错误文件恢复重试: {os.path.basename(err_file)}'
                         f'({len(failed_items)} 条)→ 回填 {os.path.basename(current_batch_file)}')

    output_filename = os.path.basename(current_batch_file)

    is_crawling = True
    current_platform = 'jd'
    crawling_task = Thread(target=run_crawl_task_from_rows,
                           args=(failed_items, current_batch_file, {'is_retry': True}))
    crawling_task.start()

    return jsonify({
        'success': True,
        'message': f'Retrying {len(failed_items)} failed items',
        'retry_count': len(failed_items),
        'output_file': output_filename
    })

@app.route('/api/crawl/stop', methods=['POST'])
def api_crawl_stop():
    """停止当前正在运行的爬取(共享接口,不区分平台)"""
    global is_crawling
    is_crawling = False
    return jsonify({'success': True, 'message': 'Crawling stopped'})

def _kill_stale_browser_processes():
    """杀掉残留的 patchright Chromium for Testing / chromedriver 进程,返回被杀的 pattern 列表.
    ⚠️ 进程命令行是 'Google Chrome for Testing'(路径里只有小写 chromium-1217),
    旧代码用 pkill -f 'Chromium'(大写 C)匹配 0 个 → 僵尸进程从不被清理,
    占着 profile 的 user-data-dir 锁,导致下次 launch_persistent_context 撞
    'Opening in existing browser session' → TargetClosedError."""
    import subprocess
    killed = []
    for pattern in ('Chrome for Testing', 'chrome_crashpad', 'chromedriver'):
        try:
            r = subprocess.run(['pkill', '-9', '-f', pattern], capture_output=True)
            if r.returncode == 0:
                killed.append(pattern)
        except Exception:
            pass
    return killed


def _row_identity(brand, item, url, product_key):
    """一行的业务身份 = (Brand, Item, URL, Product Key) 归一化元组。
    ⚠️ 不能只用 URL:源数据里允许同一个京东商品挂多条不同 Item 的行(同 URL 不同 Item),
    它们是合法的不同行,必须各自保留;只有四元组才能唯一区分。"""
    return (str(brand or '').strip(), str(item or '').strip(),
            str(url or '').strip(), str(product_key or '').strip())


def _live_row_identity(r):
    """从 live_results 行(小写键)取身份。Product Key 与写 Excel 时一致:product_key 缺失则用 product_id。"""
    return _row_identity(r.get('brand'), r.get('item'), r.get('url'),
                         r.get('product_key') or r.get('product_id'))


def _upsert_live_result(row):
    """按「完整行身份」把结果写进 live_results — 已存在同身份旧记录(retry 重爬同一行)就先删旧再加,
    避免 failed 旧行和 success 新行同时存在(否则下次 retry 会把已成功的旧 failed 行又捞出来重爬)。
    仅 retry 用;普通爬取用 append 原样保留所有行(含合法的同 URL 不同 Item 行)。"""
    global live_results
    plat = row.get('platform')
    key = _live_row_identity(row)
    live_results = [r for r in live_results
                    if not (r.get('platform') == plat and _live_row_identity(r) == key)]
    live_results.append(row)


# 错误小文件与主文件「同名配对」:主文件 JD_Price_Marks_X.xlsx ↔ 错误文件 JD_Price_Marks_X_errors.xlsx。
# 这样 Flask 重启后(live_results 已清空),retry 只需找到最新的 *_errors.xlsx,
# 就能既拿到「要重爬哪些」(读错误文件),又知道「回填到哪个主文件」(去掉 _errors 后缀)。
RETRYABLE_STATUSES = ('failed', 'blocked', 'forbidden', 'skipped')


def _errors_path_for(master_path):
    """主文件路径 → 配对的错误文件路径(……_errors.xlsx)."""
    if master_path.endswith('.xlsx'):
        return master_path[:-len('.xlsx')] + '_errors.xlsx'
    return master_path + '_errors.xlsx'


def _latest_errors_file():
    """outputs/ 下最新的京东错误文件,没有则返回 None(用于 Flask 重启后恢复 retry)."""
    import glob
    files = glob.glob(os.path.join(app.config['OUTPUT_FOLDER'], 'JD_*_errors.xlsx'))
    if not files:
        return None
    return max(files, key=os.path.getmtime)


@app.route('/api/reset_browser', methods=['POST'])
def api_reset_browser():
    """重置浏览器会话 — 关掉残留 chromium + 清空 crawler 单例.
    用户点这个按钮等价于"在终端 kill Flask 再重启"的效果,但不杀 Flask 本身.
    下次开始爬取时会重新初始化一个全新的 patchright + chromium."""
    global crawler_instance, tmall_crawler_instance, is_crawling

    if is_crawling:
        return jsonify({
            'success': False,
            'error': '当前正在爬取,请先点"停止"再重置',
        }), 400

    # 关老 context (best effort — playwright 线程可能已死,会抛 thread error,吞掉)
    for inst_name in ('crawler_instance', 'tmall_crawler_instance'):
        inst = globals().get(inst_name)
        if inst is None:
            continue
        try:
            close_fn = getattr(inst, '_close_context', None) or getattr(inst, 'close', None)
            if close_fn:
                close_fn()
        except Exception:
            pass

    crawler_instance = None
    tmall_crawler_instance = None

    # 强力清理:杀掉所有 patchright Chrome for Testing 进程,以及 chromedriver
    killed = _kill_stale_browser_processes()

    return jsonify({
        'success': True,
        'message': (
            f'浏览器会话已重置 — 下次点"开始爬取"会重新初始化一个干净的 chromium。'
            f'{"已清理: " + ", ".join(killed) if killed else ""}'
        ),
    })

# ==================== JD 账号池(profile)管理 ====================

def _free_browser_for_provisioning():
    """扫码/验证前清场:关 crawler 单例 + 杀残留 chromium,避免 profile 目录被占锁。"""
    global crawler_instance, tmall_crawler_instance
    for inst_name in ('crawler_instance', 'tmall_crawler_instance'):
        inst = globals().get(inst_name)
        if inst is None:
            continue
        try:
            close_fn = getattr(inst, '_close_context', None) or getattr(inst, 'close', None)
            if close_fn:
                close_fn()
        except Exception:
            pass
    crawler_instance = None
    tmall_crawler_instance = None
    _kill_stale_browser_processes()


def _profile_emit(event, data):
    socketio.emit(event, data)


@app.route('/api/profiles', methods=['GET'])
def api_profiles_list():
    """列出 JD 账号池槽位 + 状态(空/已登录+昵称/cooldown)。"""
    return jsonify({
        'profiles': profile_provision.list_profiles_status(),
        'next_id': profile_provision.next_free_id(),
        'busy': profile_provision.is_busy(),
        'crawling': is_crawling,
    })


@app.route('/api/profiles/scan', methods=['POST'])
def api_profiles_scan():
    """给某个 profile 扫码登录一个京东账号(后台启可见 chromium + 轮询登录态)。"""
    if is_crawling:
        return jsonify({'error': '正在爬取,请先停止再配置账号'}), 400
    data = request.json or {}
    pid = data.get('id')
    if pid is None:
        pid = profile_provision.next_free_id()
    ok, err = profile_provision.start_scan(
        int(pid), _profile_emit, before_launch=_free_browser_for_provisioning)
    if not ok:
        return jsonify({'error': err}), 400
    return jsonify({'success': True, 'id': int(pid)})


@app.route('/api/profiles/scan/cancel', methods=['POST'])
def api_profiles_scan_cancel():
    profile_provision.cancel()
    return jsonify({'success': True})


@app.route('/api/profiles/verify', methods=['POST'])
def api_profiles_verify():
    """验证某 profile 的登录是否仍有效。"""
    if is_crawling:
        return jsonify({'error': '正在爬取,请先停止再验证'}), 400
    data = request.json or {}
    pid = data.get('id')
    if pid is None:
        return jsonify({'error': '缺少 profile id'}), 400
    ok, err = profile_provision.start_verify(
        int(pid), _profile_emit, before_launch=_free_browser_for_provisioning)
    if not ok:
        return jsonify({'error': err}), 400
    return jsonify({'success': True, 'id': int(pid)})


@app.route('/api/profiles/cooldown', methods=['POST'])
def api_profiles_cooldown():
    """一键冷却/恢复某 profile(改 .cooldown 后缀,移出/移回轮换)。"""
    if is_crawling:
        return jsonify({'error': '正在爬取,请先停止再操作'}), 400
    data = request.json or {}
    pid, on = data.get('id'), bool(data.get('on'))
    if pid is None:
        return jsonify({'error': '缺少 profile id'}), 400
    ok, err = profile_provision.set_cooldown(int(pid), on)
    if not ok:
        return jsonify({'error': err}), 400
    return jsonify({'success': True})


@app.route('/api/profiles/remove', methods=['POST'])
def api_profiles_remove():
    """移除某 profile(移到 .removed 后缀,可逆)。"""
    if is_crawling:
        return jsonify({'error': '正在爬取,请先停止再操作'}), 400
    data = request.json or {}
    pid = data.get('id')
    if pid is None:
        return jsonify({'error': '缺少 profile id'}), 400
    ok, err = profile_provision.remove_profile(int(pid))
    if not ok:
        return jsonify({'error': err}), 400
    return jsonify({'success': True})


@app.route('/api/download/<filename>')
def api_download(filename):
    """下载结果文件"""
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    else:
        return jsonify({'error': 'File not found'}), 404

@app.route('/api/template')
def api_template():
    """下载上传模板 Excel"""
    wb = Workbook()

    # ----- Sheet 1: Products -----
    ws = wb.active
    ws.title = 'Products'

    headers = [
        ('BRAND', '品牌名称，如：倍思 / 安克 / 罗马仕'),
        ('Item', '型号或商品名（可留空）'),
        ('ProductUrl std', '京东商品详情页完整 URL，例如 https://item.jd.com/100140584252.html'),
        ('ProductKey', '京东商品 SKU 编号（URL 中 /xxxx.html 的数字部分）。URL 与 ProductKey 至少填一项'),
        ('Price Reference', '参考价，用于和爬取价对比（可留空）'),
    ]

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='C23A2A')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style='thin', color='D9D9D9')
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (name, comment) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border
        cell.comment = Comment(comment, 'Template')

    # 示例行
    samples = [
        ('倍思', 'PD 65W 氮化镓充电器', 'https://item.jd.com/100140584252.html', '100140584252', '199'),
        ('安克', '', 'https://item.jd.com/100157733931.html', '100157733931', ''),
        ('罗马仕', '20000mAh 移动电源', '', '100012345678', '129'),
    ]
    sample_font = Font(name='Calibri', size=11, color='8A8578', italic=True)
    for r_idx, sample in enumerate(samples, 2):
        for c_idx, value in enumerate(sample, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = sample_font
            cell.border = cell_border
            cell.alignment = Alignment(vertical='center')

    # 列宽
    widths = [10, 28, 46, 18, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    # ----- Sheet 2: 说明 -----
    ws2 = wb.create_sheet('说明')
    title_font = Font(name='Calibri', size=14, bold=True, color='2C2C2C')
    body_font = Font(name='Calibri', size=11, color='2C2C2C')
    hint_font = Font(name='Calibri', size=10, color='8A8578', italic=True)

    lines = [
        ('使用说明', title_font),
        ('', None),
        ('1. 必填字段:URL 和 ProductKey 至少填一项。两者都填时优先使用 URL。', body_font),
        ('2. ProductKey 是京东商品页 URL 中 /xxxxx.html 的数字部分,例如:', body_font),
        ('   https://item.jd.com/100140584252.html  ->  ProductKey = 100140584252', hint_font),
        ('3. BRAND / Item / Price Reference 可留空,仅用于结果对照展示。', body_font),
        ('4. 表头名称兼容写法:', body_font),
        ('   Brand / 品牌 | Item / 型号 / Model | URL / ProductUrl / 链接 | Product Key / SKU | Price Reference / 参考价', hint_font),
        ('5. 删除前 3 行示例数据后再填入自己的商品清单,保存为 .xlsx 上传即可。', body_font),
        ('', None),
        ('提示:可一次上传几十到上百行,但建议分批以降低反爬触发概率。', hint_font),
    ]
    for r_idx, (text, font) in enumerate(lines, 1):
        cell = ws2.cell(row=r_idx, column=1, value=text)
        if font:
            cell.font = font
    ws2.column_dimensions['A'].width = 90

    # 写入内存 buffer
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name='JD_Crawler_Template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/quick-check', methods=['POST'])
def api_quick_check():
    """快速查询单个商品价格"""
    global crawler_instance, is_crawling

    if is_crawling:
        return jsonify({'error': 'Crawler is busy with batch job'}), 400

    data = request.json
    product_input = data.get('product_id', '').strip()

    if not product_input:
        return jsonify({'error': 'Product ID is required'}), 400

    # Extract product ID from URL or raw ID
    match = re.search(r'(\d{6,})', product_input)
    if not match:
        return jsonify({'error': 'Invalid product ID or URL'}), 400

    product_id = match.group(1)

    # Run in a thread and return result via socket
    def do_check():
        global crawler_instance
        try:
            socketio.emit('quick_check_status', {'status': 'starting', 'product_id': product_id})

            # Initialize crawler if needed
            if not crawler_instance or not crawler_instance.is_session_valid():
                socketio.emit('quick_check_status', {'status': 'logging_in', 'product_id': product_id})
                crawler = JDCrawlerViaSearch(headless=False)
                crawler.login()
                if not crawler.is_logged_in:
                    socketio.emit('quick_check_result', {
                        'success': False,
                        'product_id': product_id,
                        'error': 'Login failed'
                    })
                    return
                crawler_instance = crawler
            else:
                crawler = crawler_instance

            socketio.emit('quick_check_status', {'status': 'fetching', 'product_id': product_id})
            prices = crawler.get_price_via_search(product_id)

            if prices:
                original = prices.get('original')
                promo = prices.get('promo')

                # Determine status
                if original in ('not_found', 'blocked', 'forbidden', 'unavailable'):
                    socketio.emit('quick_check_result', {
                        'success': False,
                        'product_id': product_id,
                        'status': original,
                        'error': f'Product status: {original}'
                    })
                else:
                    socketio.emit('quick_check_result', {
                        'success': True,
                        'product_id': product_id,
                        'original_price': original,
                        'promo_price': promo,
                        'url': f'https://item.jd.com/{product_id}.html'
                    })
            else:
                socketio.emit('quick_check_result', {
                    'success': False,
                    'product_id': product_id,
                    'error': 'Could not extract price'
                })

        except Exception as e:
            socketio.emit('quick_check_result', {
                'success': False,
                'product_id': product_id,
                'error': str(e)
            })

    thread = Thread(target=do_check)
    thread.start()

    return jsonify({'success': True, 'message': 'Quick check started', 'product_id': product_id})

# ==================== 爬取任务 ====================

def process_single_row(crawler, input_row, idx, total, batch_time):
    """处理单行并返回结果 row"""
    global is_crawling

    url = str(input_row.get('url', ''))
    # 始终从 URL 解析 product_id — 不信任 Excel 里的 ProductKey 列
    # (那列经常被污染成 "ID|商品名" 的脏数据,且本来就是另一个系统的 ID)
    m = re.search(r'/(\d+)\.html', url)
    product_id = m.group(1) if m else ''

    if not product_id:
        emit_log('WARNING', f'[{idx}/{total}] Cannot extract product ID: {url}')
        return None

    item_label = input_row.get('item') or product_id
    emit_log('INFO', f'[{idx}/{total}] Processing: {item_label}')

    # 发送进度
    emit_progress({
        'current': idx,
        'total': total,
        'percent': round(idx / total * 100, 1),
        'current_url': url,
        'product_id': product_id,
        'status': 'processing'
    })

    crawl_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = {
        'index': idx,
        'platform': 'jd',
        'brand': input_row.get('brand', ''),
        'item': input_row.get('item', ''),
        'product_key': input_row.get('product_key', ''),
        'price_reference': input_row.get('price_reference', ''),
        'product_id': product_id,
        'url': url,
        'batch_time': batch_time,
        'crawl_time': crawl_time,
        'original_price': None,
        'promo_price': None,
        'status': 'pending'
    }

    try:
        prices = crawler.get_price_via_search(product_id)

        if prices:
            original = prices.get('original')
            promo = prices.get('promo')
            diag = prices.get('_diag', '')

            if original == 'not_found':
                emit_log('WARNING', f'  Product not found')
                row.update({'status': 'not_found', 'original_price': '-', 'promo_price': '-'})
            elif original == 'blocked':
                emit_log('WARNING', f'  Anti-crawl triggered (retry) | {diag}')
                row.update({'status': 'blocked', 'original_price': '-', 'promo_price': '-'})
            elif original == 'forbidden':
                emit_log('WARNING', f'  403 Forbidden (retry) | {diag}')
                row.update({'status': 'forbidden', 'original_price': '-', 'promo_price': '-'})
            elif original == 'unavailable':
                emit_log('WARNING', f'  Product delisted')
                row.update({'status': 'unavailable', 'original_price': '-', 'promo_price': '-'})
            else:
                emit_log('INFO', f'  OK: ¥{original} / ¥{promo}')
                row.update({
                    'status': 'success',
                    'original_price': original,
                    'promo_price': promo
                })
                if not original or not promo:
                    row['status'] = 'partial'
        else:
            # 收集诊断信息(get_price 返回 None 通常是会话/网络问题,或价格元素提取失败)
            none_diag = ''
            try:
                none_diag = (f'url="{(crawler.driver.current_url or "")[:90]}" '
                             f'title="{(crawler.driver.title or "")[:50]}"')
            except Exception:
                none_diag = 'driver state unavailable'
            emit_log('WARNING', f'  Failed (can retry) | prices=None | {none_diag}')
            row.update({'status': 'failed', 'original_price': '-', 'promo_price': '-'})

    except Exception as e:
        emit_log('ERROR', f'  Error: {str(e)}')
        row.update({'status': 'failed', 'original_price': '-', 'promo_price': '-'})

    return row


def run_crawl_task(input_filepath, output_filepath, config):
    """运行爬取任务（从文件）"""
    global uploaded_urls, uploaded_rows

    emit_log('INFO', f'Reading file: {os.path.basename(input_filepath)}')
    rows = _parse_excel(input_filepath)
    uploaded_rows = rows
    uploaded_urls = [r['url'] for r in rows]
    run_crawl_task_from_rows(rows, output_filepath, config)


def run_crawl_task_from_rows(input_rows, output_filepath, config):
    """运行爬取任务(京东,从 row dict 列表)"""
    global is_crawling, current_platform, crawler_instance, current_results, live_results

    # retry 时按「行身份」替换 live_results 里的旧 failed 行(幂等);普通爬取原样 append 保留所有行
    is_retry = bool((config or {}).get('is_retry'))

    # 爬取强度:UI 传 speed=regular/fast/enhanced,决定每批条数与批间冷却(默认常规)
    preset = JD_SPEED_PRESETS.get((config or {}).get('speed'), JD_SPEED_PRESETS['regular'])
    batch_size = preset['batch_size']
    batch_cooldown = preset['cooldown']

    try:
        total = len(input_rows)
        emit_log('INFO', '=' * 50)
        emit_log('INFO', f'Starting batch crawl: {total} products')
        emit_log('INFO', '=' * 50)

        # 复用已有的浏览器实例，避免重复初始化
        if crawler_instance and crawler_instance.is_session_valid():
            emit_log('INFO', '复用已有浏览器会话')
            crawler = crawler_instance
            # 确认登录状态
            if not crawler.is_logged_in:
                emit_log('INFO', 'Logging in...')
                crawler.login()
        else:
            emit_log('INFO', 'Initializing browser...')
            # 上一轮跑完后 chromium 进程可能仍残留并占着 profile 的 user-data-dir 锁,
            # 直接 launch_persistent_context 会撞 "Opening in existing browser session"
            # → TargetClosedError(retry 失效的根因)。先关旧 context + 杀残留进程,确保干净启动。
            if crawler_instance is not None:
                try:
                    close_fn = (getattr(crawler_instance, '_close_context', None)
                                or getattr(crawler_instance, 'close', None))
                    if close_fn:
                        close_fn()
                except Exception:
                    pass
                crawler_instance = None
            killed = _kill_stale_browser_processes()
            if killed:
                emit_log('INFO', f'已清理残留浏览器进程: {", ".join(killed)}')
                time.sleep(1.5)
            crawler = JDCrawlerViaSearch(headless=False)
            crawler_instance = crawler
            emit_log('INFO', 'Logging in...')
            crawler.login()

        if not crawler.is_logged_in:
            emit_log('ERROR', 'Login failed')
            is_crawling = False
            socketio.emit('crawl_complete', {'success': False, 'error': 'Login failed'})
            return

        emit_log('INFO', 'Login successful')

        # 热身：模拟正常浏览降低反爬风险
        emit_log('INFO', '正在热身（模拟正常浏览）...')
        warmup_ok, warmup_err = crawler.warmup()

        # 热身后再做一次"真探活" — playwright 线程死了 / chromium 进程没了
        # 这种深层问题用 warmup_ok 检测不到(warmup 内部捕获后就 swallowed 了)
        if not warmup_ok or not crawler.is_session_valid():
            emit_log('ERROR', '❌ 热身失败 — 浏览器会话已失效,无法继续爬取')
            if warmup_err:
                emit_log('ERROR', f'   底层错误: {warmup_err}')
            emit_log('ERROR',
                     '   常见原因: Flask 长时间空闲后 patchright 工作线程退出,'
                     '或者上一次跑完后 chromium 进程已被关闭')
            emit_log('ERROR',
                     '   解决方法: 在终端运行 `lsof -ti :5001 | xargs kill -9` 杀掉 Flask,'
                     '然后 `python3 app.py` 重启,再点开始爬取')
            # 同时把缓存的 crawler_instance 清掉,下次启动会强制重新 init
            crawler_instance = None
            is_crawling = False
            socketio.emit('crawl_complete', {
                'success': False,
                'error': '热身失败:浏览器会话已失效,请重启 Flask 后重试',
            })
            return

        batch_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        start_time = time.time()

        success_count = 0
        failed_count = 0
        unavailable_count = 0
        consecutive_failures = 0
        anti_crawl_cooldowns = 0  # 单批内已触发反爬冷却的次数

        # 自动分批:单批 batch_size 条,批次间冷却 batch_cooldown 秒(由爬取强度预设决定)
        chunks = [input_rows[i:i + batch_size]
                  for i in range(0, total, batch_size)]
        total_batches = len(chunks)
        if total_batches > 1:
            emit_log('INFO',
                     f'自动分批({preset["label"]}强度): {total} 条 -> {total_batches} 批 '
                     f'(每批 {batch_size} 条, 批间冷却 {batch_cooldown//60} 分钟)')

        global_idx = 0
        user_stopped = False
        session_dead = False  # 浏览器会话异常死亡(被关/崩溃)—— 与"真·反爬"区分,触发后如实报告并中止
        # 一旦发现无法账号交替(只有 1 个可用账号),本次任务后续直接走冷却,
        # 不再每批反复尝试 rotate(避免对登录过期的 profile 反复 close/launch 的无谓 churn)
        single_account_mode = False

        for batch_idx, chunk in enumerate(chunks, 1):
            if not is_crawling:
                user_stopped = True
                break

            if total_batches > 1:
                emit_log('INFO', f'━━━ 第 {batch_idx}/{total_batches} 批: {len(chunk)} 条 ━━━')

            # 单批内的反爬冷却计数器重置(批与批独立)
            consecutive_failures = 0
            anti_crawl_cooldowns = 0
            # 随机游走计数器:每 10-15 条插入一次"伪浏览"(访问首页/购物车),制造行为多样性
            items_since_walk = 0
            next_walk_at = random.randint(10, 15)

            for chunk_idx, input_row in enumerate(chunk):
                if not is_crawling:
                    user_stopped = True
                    break
                global_idx += 1
                idx = global_idx

                # 触发随机游走(在请求当前商品 *之前*,让 referer 看起来像从首页/购物车点进来)
                if items_since_walk >= next_walk_at:
                    try:
                        label = crawler.random_walk()
                        emit_log('INFO', f'  ↪ 插入伪浏览: {label}(降低线性行为可疑度)')
                    except Exception as e:
                        emit_log('WARNING', f'  ↪ 伪浏览失败(忽略): {e}')
                    items_since_walk = 0
                    next_walk_at = random.randint(10, 15)

                # 单批内突发失败:先分清「浏览器会话已死(被关/崩溃)」还是「真·反爬」
                if consecutive_failures >= 2:
                    # 走 CDP 真探活。会话死了就如实中止,绝不误报成反爬去冷却重试。
                    if not crawler.is_session_valid():
                        emit_log('ERROR', '❌ 浏览器会话异常停止(可能被外部关闭或崩溃),已中止本次采集')
                        emit_log('ERROR', '   这不是反爬。请点「重置浏览器会话」后重新开始。')
                        session_dead = True
                        is_crawling = False
                        break

                    # 会话健康 → 判定为真·反爬,冷却
                    anti_crawl_cooldowns += 1
                    cooldown = min(30 + anti_crawl_cooldowns * 30, 120)
                    emit_log('WARNING', f'检测到反爬,冷却{cooldown}秒... (第{anti_crawl_cooldowns}次)')
                    time.sleep(cooldown)
                    consecutive_failures = 0

                    # 冷却后先访问京东首页"重置"会话;若此时会话已死,同样如实中止
                    try:
                        emit_log('INFO', '重置会话:访问京东首页...')
                        crawler.driver.get("https://www.jd.com")
                        time.sleep(random.uniform(3.0, 5.0))
                        crawler.driver.execute_script("window.scrollTo(0, 500);")
                        time.sleep(1)
                        crawler.driver.execute_script("window.scrollTo(0, 0);")
                        time.sleep(1)
                    except Exception:
                        if not crawler.is_session_valid():
                            emit_log('ERROR', '❌ 浏览器会话异常停止,已中止本次采集。请重置浏览器后重试')
                            session_dead = True
                            is_crawling = False
                            break

                row = process_single_row(crawler, input_row, idx, total, batch_time)
                if not row:
                    continue

                if is_retry:
                    _upsert_live_result(row)
                else:
                    live_results.append(row)
                emit_result_row(row)
                items_since_walk += 1

                if row['status'] == 'success':
                    success_count += 1
                    consecutive_failures = 0
                    anti_crawl_cooldowns = 0
                elif row['status'] in ('failed', 'blocked', 'forbidden', 'partial'):
                    failed_count += 1
                    consecutive_failures += 1
                elif row['status'] in ('unavailable', 'not_found'):
                    unavailable_count += 1

                emit_progress({
                    'statistics': {
                        'success': success_count,
                        'failed': failed_count,
                        'unavailable': unavailable_count,
                        'total': idx,
                    }
                })

                # 连续 3 次失败 — 当前 profile 被风控,切换到下一个 profile 继续
                if consecutive_failures >= 3:
                    emit_log('WARNING',
                             f'⚠ 连续 3 次失败 — 当前 profile_{crawler.current_profile_id} 可能被风控,'
                             f'尝试切换到下一个 profile...')
                    new_pid = crawler.switch_to_next_profile()
                    if new_pid is None:
                        # 所有 profile 耗尽 — 把本批剩余标记 skipped 进入下一批冷却
                        remaining_rows = chunk[chunk_idx + 1:]
                        skip_n = len(remaining_rows)
                        emit_log('ERROR',
                                 f'✗ profile 池已耗尽 — 跳过本批剩余 {skip_n} 条,'
                                 f'进入下一批冷却({batch_cooldown//60} 分钟后会重新从 profile_1 开始)')
                        crawl_time_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        for sk_row in remaining_rows:
                            global_idx += 1
                            sk_url = str(sk_row.get('url', ''))
                            m = re.search(r'/(\d+)\.html', sk_url)
                            sk_pid = m.group(1) if m else ''
                            skipped = {
                                'index': global_idx,
                                'platform': 'jd',
                                'brand': sk_row.get('brand', ''),
                                'item': sk_row.get('item', ''),
                                'product_key': sk_row.get('product_key', ''),
                                'price_reference': sk_row.get('price_reference', ''),
                                'product_id': sk_pid,
                                'url': sk_url,
                                'batch_time': batch_time,
                                'crawl_time': crawl_time_now,
                                'original_price': '-',
                                'promo_price': '-',
                                'status': 'skipped',
                            }
                            if is_retry:
                                _upsert_live_result(skipped)
                            else:
                                live_results.append(skipped)
                            emit_result_row(skipped)
                            failed_count += 1
                        emit_progress({
                            'statistics': {
                                'success': success_count,
                                'failed': failed_count,
                                'unavailable': unavailable_count,
                                'total': global_idx,
                            }
                        })
                        # 重置 profile 池游标,下一批冷却完后重新从 profile_1 开始
                        crawler.current_profile_id = None
                        break  # 跳出 chunk,进入批次间冷却

                    # 切换成功 — 重置计数器,继续当前批
                    emit_log('INFO', f'✓ 已切到 profile_{new_pid},继续爬取')
                    consecutive_failures = 0
                    anti_crawl_cooldowns = 0

                # 单条间延迟 — get_price_via_search 内部已有 10-15s 真实停留,
                # 这里只额外加少量间隔(2-4s)用于模拟"看完一个商品后切到下一个"的过渡
                time.sleep(random.uniform(2.0, 4.0))

            # 批次间:优先「账号交替」——切到下一个账号继续,用对方那批的时长填掉冷却空窗(免等)。
            # 只在两种情况下才真正冷却:① 上一批 profile 池耗尽(全员被风控,需自愈时间);
            # ② 只剩 1 个可用账号(无从交替,退回已验证的 600s)。冷却值绝不缩水。
            if batch_idx < total_batches and is_crawling:
                if crawler.current_profile_id is None:
                    # 池耗尽:给所有账号自愈时间
                    emit_log('INFO',
                             f'✓ 第 {batch_idx}/{total_batches} 批完成(账号池耗尽)— '
                             f'冷却 {batch_cooldown//60} 分钟后继续')
                    if not _batch_cooldown(batch_cooldown, platform='jd'):
                        user_stopped = True
                        break
                else:
                    new_pid = None if single_account_mode else crawler.rotate_profile()
                    if new_pid is not None:
                        emit_log('INFO',
                                 f'✓ 第 {batch_idx}/{total_batches} 批完成 — '
                                 f'切到 profile_{new_pid} 继续(账号交替,免冷却)')
                    else:
                        # 只有 1 个可用账号 → 用验证过的 600s,不缩水;并记住,后续不再尝试交替
                        single_account_mode = True
                        emit_log('INFO',
                                 f'✓ 第 {batch_idx}/{total_batches} 批完成 — '
                                 f'仅 1 个可用账号,冷却 {batch_cooldown//60} 分钟后继续')
                        if not _batch_cooldown(batch_cooldown, platform='jd'):
                            user_stopped = True
                            break

        if user_stopped and not session_dead:
            emit_log('WARNING', 'Crawl stopped by user')

        # 保存 Excel — 按「行身份」回填进磁盘上已有的主文件:
        # retry 只爬了几条、或 Flask 重启后 live_results 只剩本次几条时,主文件里「没参与本次」
        # 的其余行原样保留,本次结果就地替换对应行。用列表+身份匹配(不是按 URL 建字典),
        # 因为同 URL 不同 Item 是合法的多行,字典会把它们折叠丢行。
        emit_log('INFO', 'Saving results to Excel...')
        # 通知前端:开始生成可下载的 Excel(覆盖正常结束/停止/会话异常所有结束路径)
        socketio.emit('crawl_saving', {'platform': 'jd'})

        def _to_excel_row(r):
            return {
                'Batch Time': r['batch_time'],
                'Crawl Time': r['crawl_time'],
                'Brand': r.get('brand', ''),
                'Item': r.get('item', ''),
                'URL': r['url'],
                'Product Key': r.get('product_key', '') or r.get('product_id', ''),
                'Price Reference': r.get('price_reference', ''),
                'Status': r['status'],
                'Price': r['original_price'] if r['original_price'] not in (None, '-') else 'N/A',
                'Promotion Price': r['promo_price'] if r['promo_price'] not in (None, '-') else 'N/A',
            }

        # 本次会话的京东结果,按行身份建索引(同身份取最后一条)
        session_by_id = {}
        for r in live_results:
            if r.get('platform') == 'jd' and r.get('url'):
                session_by_id[_live_row_identity(r)] = _to_excel_row(r)

        out_rows = []
        if os.path.exists(output_filepath):
            # 主文件已存在(retry/重启):逐行读旧主文件,身份命中本次结果就就地替换,否则原样保留 —— 保序、保数量
            used = set()
            try:
                old_df = pd.read_excel(output_filepath)
                for _, orow in old_df.iterrows():
                    oid = _row_identity(orow.get('Brand'), orow.get('Item'),
                                        orow.get('URL'), orow.get('Product Key'))
                    if oid in session_by_id:
                        out_rows.append(session_by_id[oid])
                        used.add(oid)
                    else:
                        out_rows.append({c: ('' if pd.isna(orow[c]) else orow[c])
                                         for c in old_df.columns})
            except Exception as e:
                emit_log('WARNING', f'读取旧主文件失败(忽略,按本次结果全新写入): {e}')
                out_rows = []
                used = set()
            # 本次有、但旧主文件里没有的(全新行)追加到末尾
            for sid, srow in session_by_id.items():
                if sid not in used:
                    out_rows.append(srow)
        else:
            # 全新一批:原样写出本次所有行(普通爬取用 append,合法的同 URL 不同 Item 行都在)
            out_rows = [_to_excel_row(r) for r in live_results
                        if r.get('platform') == 'jd' and r.get('url')]

        df_results = pd.DataFrame(out_rows)
        df_results.to_excel(output_filepath, index=False, engine='openpyxl')

        # 产出/更新独立错误小文件(上传格式 5 列)——可持久化、可手动当新批次重跑、retry 的数据源。
        #    从合并后的主文件真相派生「当前还失败的」,全成功则删掉旧错误文件避免误导。
        errors_path = _errors_path_for(output_filepath)
        err_rows = [row for row in out_rows
                    if str(row.get('Status', '')) in RETRYABLE_STATUSES]
        if err_rows:
            err_df = pd.DataFrame([{
                'Brand': row.get('Brand', ''),
                'Item': row.get('Item', ''),
                'URL': row.get('URL', ''),
                'Product Key': row.get('Product Key', ''),
                'Price Reference': row.get('Price Reference', ''),
            } for row in err_rows])
            err_df.to_excel(errors_path, index=False, engine='openpyxl')
            errors_file_name = os.path.basename(errors_path)
            emit_log('INFO',
                     f'  ⚠ {len(err_rows)} 条失败/跳过 → 错误文件 {errors_file_name}'
                     f'(可点「重试失败项」,或手动当新批次重新上传)')
        else:
            errors_file_name = None
            try:
                if os.path.exists(errors_path):
                    os.remove(errors_path)
            except Exception:
                pass

        duration = time.time() - start_time
        emit_log('INFO', '=' * 50)
        emit_log('INFO', f'Crawl complete!')
        emit_log('INFO', f'  Success: {success_count}')
        emit_log('INFO', f'  Failed: {failed_count}')
        emit_log('INFO', f'  Unavailable: {unavailable_count}')
        emit_log('INFO', f'  Duration: {duration:.1f}s')
        emit_log('INFO', f'  Output: {os.path.basename(output_filepath)}')
        emit_log('INFO', '=' * 50)

        if session_dead:
            emit_log('ERROR', f'本次因浏览器会话异常而中止 —— 已采集的 {success_count} 条结果已保存,可直接下载')
            socketio.emit('crawl_complete', {
                'success': False,
                'platform': 'jd',
                'error': '浏览器会话异常停止(被关闭或崩溃),已中止。已采结果可下载;请「重置浏览器会话」后重试剩余。',
                # 带上产出文件名,让前端照常给出「下载结果 / 错误记录 / 重试」入口
                'output_file': os.path.basename(output_filepath),
                'errors_file': errors_file_name,
            })
            is_crawling = False
            current_platform = None
            current_results = live_results
            return

        socketio.emit('crawl_complete', {
            'success': True,
            'platform': 'jd',
            'output_file': os.path.basename(output_filepath),
            'errors_file': errors_file_name,
            'stats': {
                'success': success_count,
                'failed': failed_count,
                'unavailable': unavailable_count,
                'total': sum(1 for r in live_results if r.get('platform') == 'jd'),
                'duration': round(duration, 1)
            }
        })

        # 不关闭浏览器，下次复用（避免重新下载 ChromeDriver）
        is_crawling = False
        current_platform = None
        current_results = live_results

    except Exception as e:
        emit_log('ERROR', f'Crawl task error: {str(e)}')
        import traceback
        traceback.print_exc()
        is_crawling = False
        current_platform = None

        socketio.emit('crawl_complete', {
            'success': False,
            'platform': 'jd',
            'error': str(e)
        })

# ==================== 天猫路由 ====================

@app.route('/api/tmall/upload', methods=['POST'])
def api_tmall_upload():
    """上传天猫 Excel 文件"""
    global uploaded_tmall_rows

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Invalid file type'}), 400

    filename = secure_filename(file.filename)
    # 加 tmall_ 前缀避免和京东上传文件冲突
    save_name = f'tmall_{filename}'
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], save_name)
    file.save(filepath)

    try:
        df = pd.read_excel(filepath)
        rows = _parse_tmall_excel(filepath)
        uploaded_tmall_rows = rows

        # 校验:URL 不是天猫/淘宝域名时给警告
        bad_urls = []
        for i, r in enumerate(rows, 1):
            url = r.get('url', '')
            if url and not any(d in url for d in ['tmall.com', 'taobao.com']):
                bad_urls.append({'row': i, 'url': url})

        return jsonify({
            'success': True,
            'filename': filename,
            'filepath': filepath,
            'url_count': len(rows),
            'columns': list(df.columns),
            'warnings': {'non_tmall_urls': bad_urls} if bad_urls else {}
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/tmall/preview')
def api_tmall_preview():
    """预览天猫上传数据"""
    global uploaded_tmall_rows

    if not uploaded_tmall_rows:
        return jsonify({'error': 'No file uploaded'}), 400

    rows = []
    for i, r in enumerate(uploaded_tmall_rows):
        rows.append({
            'index': i + 1,
            'brand': r.get('brand', ''),
            'item': r.get('item', ''),
            'url': r.get('url', ''),
            'product_key': r.get('product_key', ''),
            'price_reference': r.get('price_reference', ''),
        })
    return jsonify({'success': True, 'rows': rows, 'total': len(rows)})


@app.route('/api/tmall/template')
def api_tmall_template():
    """下载天猫上传模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'

    headers = [
        ('BRAND', '品牌名称,如:倍思 / 安克 / 罗马仕'),
        ('Item', '型号或商品名(可留空)'),
        ('ProductUrl tmall', '【必填】天猫/淘宝商品详情页完整 URL,例如 https://detail.tmall.com/item.htm?id=634015385434'),
        ('ProductKey', '业务侧编号(可留空,只用于显示和对照,不参与爬取)'),
        ('Price Reference', '参考价,用于和爬取价对比(可留空)'),
    ]

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='C23A2A')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style='thin', color='D9D9D9')
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (name, comment) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border
        cell.comment = Comment(comment, 'Template')

    samples = [
        ('倍思', 'PD 充电宝 20000mAh', 'https://detail.tmall.com/item.htm?id=634015385434', 'SKU-001', '199'),
        ('icon', '便携充电宝', 'https://detail.tmall.com/item.htm?id=817030084492', 'SKU-002', ''),
        ('小米', '充电宝', 'https://detail.tmall.com/item.htm?id=864497080438', 'SKU-003', '149'),
    ]
    sample_font = Font(name='Calibri', size=11, color='8A8578', italic=True)
    for r_idx, sample in enumerate(samples, 2):
        for c_idx, value in enumerate(sample, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = sample_font
            cell.border = cell_border
            cell.alignment = Alignment(vertical='center')

    widths = [10, 28, 50, 18, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    # 说明页
    ws2 = wb.create_sheet('说明')
    title_font = Font(name='Calibri', size=14, bold=True, color='2C2C2C')
    body_font = Font(name='Calibri', size=11, color='2C2C2C')
    hint_font = Font(name='Calibri', size=10, color='8A8578', italic=True)

    lines = [
        ('天猫/淘宝模板使用说明', title_font),
        ('', None),
        ('1. 【必填】ProductUrl tmall — 天猫/淘宝商品详情页完整 URL。系统会自动从 URL 提取真实商品 ID。', body_font),
        ('   例如:https://detail.tmall.com/item.htm?id=634015385434', hint_font),
        ('2. ProductKey:你的业务侧编号(SKU / 内部产品编码等),只用于显示和对照,不参与爬取,可留空。', body_font),
        ('3. BRAND / Item / Price Reference 可留空,仅用于结果对照展示。', body_font),
        ('4. 表头名称兼容写法:', body_font),
        ('   BRAND / Brand / 品牌 | Item / 型号 | ProductUrl tmall / URL / 链接 | Product Key / ProductKey / SKU | Price Reference', hint_font),
        ('5. 删除前 3 行示例数据后再填入自己的商品清单,保存为 .xlsx 上传。', body_font),
        ('', None),
        ('提示:', hint_font),
        ('- 天猫反爬强:每次批量爬取开始时需在浏览器中扫码登录(约 10 秒)。', hint_font),
        ('- 单批 30 个以内为宜,避免触发风控。', hint_font),
        ('- 请确保 URL 是天猫(detail.tmall.com)或淘宝(item.taobao.com)真实可访问的商品页。', hint_font),
    ]
    for r_idx, (text, font) in enumerate(lines, 1):
        cell = ws2.cell(row=r_idx, column=1, value=text)
        if font:
            cell.font = font
    ws2.column_dimensions['A'].width = 90

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='Tmall_Crawler_Template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/tmall/crawl/start', methods=['POST'])
def api_tmall_crawl_start():
    """开始批量爬取(天猫)"""
    global is_crawling, current_platform, current_tmall_batch_file, live_results

    if is_crawling:
        return jsonify({'error': f'Crawler is already running ({current_platform})'}), 400

    data = request.json or {}
    filepath = data.get('filepath')

    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'Invalid file path'}), 400

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"Tmall_Price_{timestamp}.xlsx"
    current_tmall_batch_file = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)

    # 清掉旧的天猫结果(保留京东的)
    live_results = [r for r in live_results if r.get('platform') != 'tmall']

    is_crawling = True
    current_platform = 'tmall'
    task = Thread(target=run_tmall_crawl_task, args=(filepath, current_tmall_batch_file))
    task.start()

    return jsonify({
        'success': True,
        'message': 'Tmall crawling started',
        'output_file': output_filename
    })


@app.route('/api/tmall/crawl/retry', methods=['POST'])
def api_tmall_crawl_retry():
    """重试失败的天猫商品"""
    global is_crawling, current_platform, current_tmall_batch_file, live_results

    if is_crawling:
        return jsonify({'error': f'Crawler is already running ({current_platform})'}), 400

    failed_items = [r for r in live_results
                    if r.get('platform') == 'tmall'
                    and r.get('status') in ('failed', 'blocked', 'no_price')]

    if not failed_items:
        return jsonify({'error': 'No failed Tmall items to retry'}), 400

    # 转回 input row 格式
    retry_rows = [{
        'brand': r.get('brand', ''),
        'item': r.get('item', ''),
        'url': r.get('url', ''),
        'item_id': r.get('item_id') or r.get('product_id', ''),
        'product_key': r.get('product_key', ''),
        'price_reference': r.get('price_reference', ''),
    } for r in failed_items]

    # 移除待重试项
    live_results = [r for r in live_results
                    if not (r.get('platform') == 'tmall'
                            and r.get('status') in ('failed', 'blocked', 'no_price'))]

    # 复用原文件名 — retry 后覆盖,得到完整最终版
    if not current_tmall_batch_file or not os.path.exists(os.path.dirname(current_tmall_batch_file) or '.'):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        current_tmall_batch_file = os.path.join(app.config['OUTPUT_FOLDER'], f"Tmall_Price_{timestamp}.xlsx")

    output_filename = os.path.basename(current_tmall_batch_file)

    is_crawling = True
    current_platform = 'tmall'
    task = Thread(target=run_tmall_crawl_task_from_rows, args=(retry_rows, current_tmall_batch_file))
    task.start()

    return jsonify({
        'success': True,
        'message': f'Retrying {len(retry_rows)} failed Tmall items',
        'retry_count': len(retry_rows),
        'output_file': output_filename
    })


# ==================== 天猫爬取任务 ====================

def _process_tmall_row(crawler, input_row, idx, total, batch_time):
    """处理单条天猫行.
    input_row 来自 _parse_tmall_excel,字段:
        url(必填,爬取依据), item_id(从 URL 提取的真实 tmall id),
        product_key(业务侧编号,只显示)
    """
    url = input_row.get('url', '')
    item_id = input_row.get('item_id', '')  # 真实 tmall id,爬取用

    if not item_id:
        emit_log('WARNING', f'[{idx}/{total}] 无法从 URL 提取商品 ID,跳过: {url}', platform='tmall')
        return None

    item_label = input_row.get('item') or input_row.get('product_key') or item_id
    emit_log('INFO', f'[{idx}/{total}] 处理: {item_label}', platform='tmall')

    emit_progress({
        'current': idx,
        'total': total,
        'percent': round(idx / total * 100, 1),
        'current_url': url,
        'product_id': item_id,
        'status': 'processing',
        'platform': 'tmall',
    })

    crawl_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = {
        'index': idx,
        'platform': 'tmall',
        'brand': input_row.get('brand', ''),
        'item': input_row.get('item', ''),
        'item_id': item_id,  # 真实 tmall id
        'product_id': item_id,
        'product_key': input_row.get('product_key', ''),  # 业务侧编号
        'price_reference': input_row.get('price_reference', ''),
        'url': url,
        'batch_time': batch_time,
        'crawl_time': crawl_time,
        'original_price': None,
        'promo_price': None,
        'shop': None,
        'title': None,
        'status': 'pending',
    }

    try:
        result = crawler.get_price(item_id)
        if result:
            original = result.get('original')
            promo = result.get('promo')
            diag = result.get('_diag')

            if original == 'blocked':
                emit_log('WARNING', f'  被拦截(可重试) | {diag or ""}', platform='tmall')
                row.update({'status': 'blocked', 'original_price': '-', 'promo_price': '-'})
            elif original == 'slider':
                emit_log('WARNING',
                         f'  ⚠️ 商品页被滑块拦截,请在浏览器手动拖动滑块完成验证后点"重试失败项" | {diag or ""}',
                         platform='tmall')
                row.update({'status': 'blocked', 'original_price': '-', 'promo_price': '-'})
            elif original == 'not_found':
                emit_log('WARNING', '  商品不存在', platform='tmall')
                row.update({'status': 'not_found', 'original_price': '-', 'promo_price': '-'})
            elif not original and not promo:
                emit_log('WARNING', f'  抽取失败(可重试) | {diag or ""}', platform='tmall')
                row.update({'status': 'no_price', 'original_price': '-', 'promo_price': '-'})
            else:
                emit_log('INFO', f'  ✓ ¥{original} / ¥{promo}', platform='tmall')
                row.update({
                    'status': 'success',
                    'original_price': original,
                    'promo_price': promo,
                    'shop': result.get('shop'),
                    'title': result.get('title'),
                    'original_label': result.get('original_label'),
                    'promo_label': result.get('promo_label'),
                })
                if not original or not promo:
                    row['status'] = 'partial'
        else:
            emit_log('WARNING', '  失败(可重试)', platform='tmall')
            row.update({'status': 'failed', 'original_price': '-', 'promo_price': '-'})

    except Exception as e:
        emit_log('ERROR', f'  错误: {str(e)}', platform='tmall')
        row.update({'status': 'failed', 'original_price': '-', 'promo_price': '-'})

    return row


def run_tmall_crawl_task(input_filepath, output_filepath):
    """运行天猫爬取(从文件)"""
    global uploaded_tmall_rows

    emit_log('INFO', f'读取文件: {os.path.basename(input_filepath)}', platform='tmall')
    rows = _parse_tmall_excel(input_filepath)
    uploaded_tmall_rows = rows
    run_tmall_crawl_task_from_rows(rows, output_filepath)


def run_tmall_crawl_task_from_rows(input_rows, output_filepath):
    """运行天猫爬取(从 row dict 列表)"""
    global is_crawling, current_platform, tmall_crawler_instance, live_results

    try:
        total = len(input_rows)
        emit_log('INFO', '=' * 50, platform='tmall')
        emit_log('INFO', f'开始批量爬取: {total} 个商品', platform='tmall')
        emit_log('INFO', '=' * 50, platform='tmall')

        # 复用浏览器实例
        if tmall_crawler_instance and tmall_crawler_instance.is_session_valid() and tmall_crawler_instance.is_logged_in:
            emit_log('INFO', '复用已有浏览器会话', platform='tmall')
            crawler = tmall_crawler_instance
        else:
            emit_log('INFO', '初始化天猫浏览器...', platform='tmall')
            crawler = TmallCrawler(headless=False)
            tmall_crawler_instance = crawler
            emit_log('INFO', '请在浏览器中扫码登录(如果先看到滑动验证,请拖动滑块完成验证后再扫码)...', platform='tmall')
            # 把滑块提示通过 socket 推到前端
            crawler.login(slider_callback=lambda msg: emit_log('WARNING', msg, platform='tmall'))

        if not crawler.is_logged_in:
            emit_log('ERROR', '登录失败,中止爬取', platform='tmall')
            is_crawling = False
            current_platform = None
            socketio.emit('crawl_complete', {'success': False, 'platform': 'tmall', 'error': 'Login failed'})
            return

        emit_log('INFO', '✓ 登录成功', platform='tmall')

        batch_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        start_time = time.time()

        success_count = 0
        failed_count = 0
        unavailable_count = 0

        # 自动分批:单批 TMALL_BATCH_SIZE 条,批次间冷却 TMALL_BATCH_COOLDOWN 秒
        chunks = [input_rows[i:i + TMALL_BATCH_SIZE]
                  for i in range(0, total, TMALL_BATCH_SIZE)]
        total_batches = len(chunks)
        if total_batches > 1:
            emit_log('INFO',
                     f'自动分批: {total} 条 -> {total_batches} 批 '
                     f'(每批 {TMALL_BATCH_SIZE} 条, 批间冷却 {TMALL_BATCH_COOLDOWN//60} 分钟)',
                     platform='tmall')

        global_idx = 0
        user_stopped = False

        for batch_idx, chunk in enumerate(chunks, 1):
            if not is_crawling:
                user_stopped = True
                break

            if total_batches > 1:
                emit_log('INFO', f'━━━ 第 {batch_idx}/{total_batches} 批: {len(chunk)} 条 ━━━', platform='tmall')

            for input_row in chunk:
                if not is_crawling:
                    user_stopped = True
                    break
                global_idx += 1
                idx = global_idx

                row = _process_tmall_row(crawler, input_row, idx, total, batch_time)
                if not row:
                    continue

                live_results.append(row)
                emit_result_row(row)

                if row['status'] == 'success':
                    success_count += 1
                elif row['status'] in ('failed', 'blocked', 'no_price', 'partial'):
                    failed_count += 1
                elif row['status'] in ('not_found',):
                    unavailable_count += 1

                emit_progress({
                    'statistics': {
                        'success': success_count,
                        'failed': failed_count,
                        'unavailable': unavailable_count,
                        'total': idx,
                    },
                    'platform': 'tmall',
                })

                # 单条间延迟(天猫比京东更保守)
                if idx <= 10:
                    delay = random.uniform(3.0, 5.0)
                elif idx <= 20:
                    delay = random.uniform(5.0, 8.0)
                else:
                    delay = random.uniform(7.0, 12.0)
                time.sleep(delay)

            # 批次间冷却(最后一批跳过)
            if batch_idx < total_batches and is_crawling:
                emit_log('INFO',
                         f'✓ 第 {batch_idx}/{total_batches} 批完成 — '
                         f'冷却 {TMALL_BATCH_COOLDOWN//60} 分钟后继续下一批',
                         platform='tmall')
                if not _batch_cooldown(TMALL_BATCH_COOLDOWN, platform='tmall'):
                    user_stopped = True
                    break

        if user_stopped:
            emit_log('WARNING', '用户中止爬取', platform='tmall')

        # 保存 Excel
        emit_log('INFO', '保存结果到 Excel...', platform='tmall')
        tmall_rows_in_results = [r for r in live_results if r.get('platform') == 'tmall']
        excel_rows = []
        for r in tmall_rows_in_results:
            excel_rows.append({
                'Batch Time': r.get('batch_time', ''),
                'Crawl Time': r.get('crawl_time', ''),
                'Brand': r.get('brand', ''),
                'Item': r.get('item', ''),
                'Shop': r.get('shop', ''),
                'URL': r.get('url', ''),
                'item_id': r.get('item_id', ''),
                'Price Reference': r.get('price_reference', ''),
                'Status': r.get('status', ''),
                'Original Price': r.get('original_price') if r.get('original_price') not in (None, '-') else 'N/A',
                'Promo Price': r.get('promo_price') if r.get('promo_price') not in (None, '-') else 'N/A',
                'Title': r.get('title', ''),
            })
        df_results = pd.DataFrame(excel_rows)
        df_results.to_excel(output_filepath, index=False, engine='openpyxl')

        duration = time.time() - start_time
        emit_log('INFO', '=' * 50, platform='tmall')
        emit_log('INFO', f'爬取完成!  成功: {success_count}  失败: {failed_count}  下架: {unavailable_count}', platform='tmall')
        emit_log('INFO', f'  用时: {duration:.1f}s  输出: {os.path.basename(output_filepath)}', platform='tmall')
        emit_log('INFO', '=' * 50, platform='tmall')

        socketio.emit('crawl_complete', {
            'success': True,
            'platform': 'tmall',
            'output_file': os.path.basename(output_filepath),
            'stats': {
                'success': success_count,
                'failed': failed_count,
                'unavailable': unavailable_count,
                'total': len(tmall_rows_in_results),
                'duration': round(duration, 1),
            }
        })

        is_crawling = False
        current_platform = None

    except Exception as e:
        emit_log('ERROR', f'天猫爬取任务异常: {str(e)}', platform='tmall')
        import traceback
        traceback.print_exc()
        is_crawling = False
        current_platform = None
        socketio.emit('crawl_complete', {
            'success': False,
            'platform': 'tmall',
            'error': str(e),
        })


# ==================== Socket.IO事件 ====================

@socketio.on('connect')
def handle_connect():
    """客户端连接"""
    print('Client connected')
    socketio.emit('connected', {'data': 'Connected to server'})

@socketio.on('disconnect')
def handle_disconnect():
    """客户端断开"""
    print('Client disconnected')

# ==================== 启动应用 ====================

if __name__ == '__main__':
    print("=" * 70)
    print("JD Price Crawler")
    print("=" * 70)
    print("\nOpen: http://localhost:5001")
    print("\nPress Ctrl+C to stop\n")

    socketio.run(app, debug=False, host='0.0.0.0', port=5001, allow_unsafe_werkzeug=True)
