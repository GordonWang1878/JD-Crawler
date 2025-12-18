#!/usr/bin/env python3
"""
京东价格批量爬取工具
支持原价和促销价双价格提取
使用搜索方式绕过反爬检测
"""
import warnings
# 抑制 urllib3 的 OpenSSL 警告
warnings.filterwarnings('ignore', message='urllib3 v2 only supports OpenSSL 1.1.1+')

import pandas as pd
from datetime import datetime
import time
import random
import os
import re
from jd_crawler_via_search import JDCrawlerViaSearch


def main():
    """主函数"""
    print("=" * 70)
    print("京东价格批量爬取工具 - 双价格版本")
    print("=" * 70)

    # 文件路径
    input_file = "Product URL List.xlsx"
    output_file = "Price Marks.xlsx"
    sheet_name = "JD Top Model by Brand"

    # 检查输入文件
    if not os.path.exists(input_file):
        print(f"✗ 找不到文件: {input_file}")
        return

    # 读取URL列表
    print(f"\n正在读取 {input_file}...")
    try:
        df_urls = pd.read_excel(input_file, sheet_name=sheet_name)
        print(f"✓ 成功读取 {len(df_urls)} 条记录")
    except Exception as e:
        print(f"✗ 读取Excel失败: {str(e)}")
        return

    # 获取URL列表
    if 'URL' not in df_urls.columns:
        print("✗ 未找到'URL'列")
        return

    urls = df_urls['URL'].dropna().tolist()
    print(f"  找到 {len(urls)} 个有效URL")

    # 询问用户运行模式
    print("\n" + "-" * 70)
    mode = input("选择运行模式:\n  1. 测试模式（只爬取前3个URL）\n  2. 小批量模式（爬取前10个URL）\n  3. 完整模式（爬取全部URL）\n请输入 1、2 或 3: ").strip()

    if mode == "1":
        urls = urls[:3]
        print(f"\n✓ 测试模式：将爬取前3个URL")
    elif mode == "2":
        urls = urls[:10]
        print(f"\n✓ 小批量模式：将爬取前10个URL")
    else:
        print(f"\n✓ 完整模式：将爬取全部{len(urls)}个URL")
        confirm = input(f"这将花费约 {len(urls) * 5 / 60:.0f}-{len(urls) * 10 / 60:.0f} 分钟，确认继续？(y/n): ").strip().lower()
        if confirm != 'y':
            print("已取消")
            return

    print("\n" + "-" * 70)

    # 初始化爬虫（有窗口模式，方便观察）
    print("\n初始化浏览器...")
    crawler = JDCrawlerViaSearch(headless=False)

    try:
        # 登录（会自动处理 cookies 失效的情况）
        crawler.login()

        if not crawler.is_logged_in:
            print("\n✗ 登录失败")
            return

        print("✓ 登录成功！")

        # 记录批次开始时间
        batch_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        batch_start_timestamp = time.time()

        # 开始爬取
        results = []
        success_count = 0
        failed_count = 0
        unavailable_count = 0  # 已下架的商品
        total_item_time = 0  # 累计处理时间

        print("\n" + "=" * 70)
        print("开始爬取价格")
        print("=" * 70 + "\n")

        for idx, url in enumerate(urls, 1):
            item_start_time = time.time()
            print(f"[{idx}/{len(urls)}] {url}")

            # 定期重启浏览器（每50个商品），避免内存泄漏
            if idx > 1 and (idx - 1) % 50 == 0:
                print(f"\n  🔄 已完成 {idx-1} 个商品，重启浏览器释放内存...\n")
                crawler.restart_browser()
                time.sleep(1.5)  # 优化后：3s → 2s → 1.5s

            # 提取商品ID
            match = re.search(r'/(\d+)\.html', url)
            if not match:
                print(f"  ✗ 无法提取商品ID\n")
                crawl_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                results.append({
                    'Batch Time': batch_time,
                    'Crawl Time': crawl_time,
                    'URL': url,
                    'Price': 'N/A',
                    'Promotion Price': 'N/A'
                })
                failed_count += 1
                continue

            product_id = match.group(1)

            # 获取价格（带重试机制）
            max_retries = 2
            retry_count = 0
            prices = None

            while retry_count <= max_retries:
                try:
                    # 检查会话是否有效
                    if not crawler.is_session_valid():
                        print(f"  ⚠️  会话失效，尝试重启浏览器...")
                        if not crawler.restart_browser():
                            print(f"  ✗ 浏览器重启失败")
                            break
                        time.sleep(1.0)  # 优化后：2s → 1.5s → 1s

                    prices = crawler.get_price_via_search(product_id)
                    break  # 成功则退出重试循环

                except Exception as e:
                    error_msg = str(e)
                    if "invalid session id" in error_msg.lower():
                        retry_count += 1
                        if retry_count <= max_retries:
                            print(f"  ⚠️  会话失效，第 {retry_count} 次重试...")
                            if crawler.restart_browser():
                                time.sleep(1.0)  # 优化后：2s → 1.5s → 1s
                                continue
                            else:
                                print(f"  ✗ 浏览器重启失败")
                                break
                        else:
                            print(f"  ✗ 重试 {max_retries} 次后仍失败")
                            break
                    else:
                        # 其他错误直接抛出
                        raise

            # 处理结果
            try:

                if prices:
                    original = prices.get('original')
                    promo = prices.get('promo')

                    # 检查商品是否下架
                    if original == 'unavailable' and promo == 'unavailable':
                        print(f"  ⚠️  商品已下架")
                        crawl_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        results.append({
                            'Batch Time': batch_time,
                            'Crawl Time': crawl_time,
                            'URL': url,
                            'Price': 'Unavailable',
                            'Promotion Price': 'Unavailable'
                        })
                        unavailable_count += 1
                        continue

                    # 检查商品是否不存在（真实无法访问）
                    if original == 'not_found' and promo == 'not_found':
                        print(f"  ⚠️  商品不存在")
                        crawl_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        results.append({
                            'Batch Time': batch_time,
                            'Crawl Time': crawl_time,
                            'URL': url,
                            'Price': 'Not Found',
                            'Promotion Price': 'Not Found'
                        })
                        unavailable_count += 1
                        continue

                    # 检查是否触发反爬验证（可重试）
                    if original == 'blocked' and promo == 'blocked':
                        print(f"  ⚠️  触发反爬验证 (建议重试)")
                        crawl_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        results.append({
                            'Batch Time': batch_time,
                            'Crawl Time': crawl_time,
                            'URL': url,
                            'Price': 'Blocked (Retry)',
                            'Promotion Price': 'Blocked (Retry)'
                        })
                        failed_count += 1
                        continue

                    # 检查是否403禁止访问（反爬拦截，可重试）
                    if original == 'forbidden' and promo == 'forbidden':
                        print(f"  ⚠️  403禁止访问 (建议重试)")
                        crawl_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        results.append({
                            'Batch Time': batch_time,
                            'Crawl Time': crawl_time,
                            'URL': url,
                            'Price': 'Forbidden (Retry)',
                            'Promotion Price': 'Forbidden (Retry)'
                        })
                        failed_count += 1
                        continue

                    # 安全检查：确保至少有一个价格，且两个字段都有值
                    # 商品一定有常规价格，可能没有促销价
                    if original and promo:
                        # 最理想的情况：两个价格都有
                        print(f"  ✓ 原价: ¥{original}, 促销价: ¥{promo}")
                        success_count += 1
                    elif original and not promo:
                        # 只有原价，说明无促销，两个字段写相同价格
                        print(f"  ✓ 原价: ¥{original} (无促销)")
                        promo = original
                        success_count += 1
                    elif promo and not original:
                        # 只找到促销价，实际上这应该是常规价格
                        print(f"  ✓ 价格: ¥{promo} (作为原价和促销价)")
                        original = promo
                        success_count += 1
                    else:
                        print(f"  ✗ 未找到价格 (可重试)")
                        failed_count += 1

                    # 保存结果（现在 original 和 promo 要么都有值，要么都是 None）
                    crawl_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    results.append({
                        'Batch Time': batch_time,
                        'Crawl Time': crawl_time,
                        'URL': url,
                        'Price': original if original else 'N/A (Retry)',
                        'Promotion Price': promo if promo else 'N/A (Retry)'
                    })
                else:
                    print(f"  ✗ 获取失败 (可重试)")
                    failed_count += 1
                    crawl_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    results.append({
                        'Batch Time': batch_time,
                        'Crawl Time': crawl_time,
                        'URL': url,
                        'Price': 'N/A (Retry)',
                        'Promotion Price': 'N/A (Retry)'
                    })

            except Exception as e:
                print(f"  ✗ 错误: {e} (可重试)")
                failed_count += 1
                crawl_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                results.append({
                    'Batch Time': batch_time,
                    'Crawl Time': crawl_time,
                    'URL': url,
                    'Price': 'N/A (Retry)',
                    'Promotion Price': 'N/A (Retry)'
                })

            # 计算和显示时间信息
            item_elapsed = time.time() - item_start_time
            total_item_time += item_elapsed
            avg_time_per_item = total_item_time / idx
            remaining_items = len(urls) - idx
            estimated_remaining_time = avg_time_per_item * remaining_items

            print(f"  ⏱️  本商品用时: {item_elapsed:.1f}秒 | 平均: {avg_time_per_item:.1f}秒/商品", end="")
            if remaining_items > 0:
                minutes = int(estimated_remaining_time // 60)
                seconds = int(estimated_remaining_time % 60)
                print(f" | 预计剩余: {minutes}分{seconds}秒")
            else:
                print()

            # 添加延迟
            if idx < len(urls):
                delay = random.uniform(1.5, 2.5)  # 优化后：3-5s → 2-3s → 1.5-2.5s
                print(f"  等待 {delay:.1f} 秒...\n")
                time.sleep(delay)

        # 显示统计
        total_elapsed = time.time() - batch_start_timestamp
        total_minutes = int(total_elapsed // 60)
        total_seconds = int(total_elapsed % 60)

        print("\n" + "=" * 70)
        print("爬取完成！")
        print("=" * 70)
        print(f"  成功: {success_count}")
        print(f"  已下架: {unavailable_count}")
        print(f"  失败: {failed_count}")
        print(f"  总计: {len(urls)}")
        if len(urls) > 0:
            print(f"  成功率: {success_count/len(urls)*100:.1f}%")
            if unavailable_count > 0:
                print(f"  下架率: {unavailable_count/len(urls)*100:.1f}%")
        print(f"\n  ⏱️  总用时: {total_minutes}分{total_seconds}秒")
        if len(urls) > 0:
            avg_per_item = total_elapsed / len(urls)
            print(f"  平均: {avg_per_item:.1f}秒/商品")

        # 保存结果
        print(f"\n正在保存结果到 {output_file}...")

        try:
            from openpyxl import load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows

            # 创建新数据
            df_new = pd.DataFrame(results)

            if os.path.exists(output_file):
                # 文件存在，追加数据（保留 Excel Table 格式）
                print(f"  检测到现有文件，追加新数据...")

                # 加载现有工作簿
                wb = load_workbook(output_file)

                # 获取 Marks sheet
                if 'Marks' in wb.sheetnames:
                    ws = wb['Marks']

                    # 检查是否需要迁移旧格式（Runtime → Batch Time + Crawl Time）
                    header_row = [cell.value for cell in ws[1]]

                    if 'Runtime' in header_row and 'Batch Time' not in header_row:
                        print(f"  检测到旧格式 (Runtime 列)，自动迁移...")

                        # 找到 Runtime 列的位置
                        runtime_col_idx = header_row.index('Runtime') + 1  # openpyxl 从1开始

                        # 1. 将 "Runtime" 改为 "Batch Time"
                        ws.cell(row=1, column=runtime_col_idx, value='Batch Time')
                        print(f"    ✓ 'Runtime' → 'Batch Time'")

                        # 2. 在 Batch Time 右边插入新列 "Crawl Time"
                        ws.insert_cols(runtime_col_idx + 1)
                        ws.cell(row=1, column=runtime_col_idx + 1, value='Crawl Time')
                        print(f"    ✓ 插入 'Crawl Time' 列")

                        # 3. 复制 Batch Time 的值到 Crawl Time（旧数据没有分开记录）
                        for row_idx in range(2, ws.max_row + 1):
                            batch_time_value = ws.cell(row=row_idx, column=runtime_col_idx).value
                            ws.cell(row=row_idx, column=runtime_col_idx + 1, value=batch_time_value)
                        print(f"    ✓ 已填充 {ws.max_row - 1} 行数据的 Crawl Time")

                        # 4. 如果有 Table，更新其范围（多了一列）
                        if ws.tables:
                            table_name = list(ws.tables.keys())[0]
                            table = ws.tables[table_name]
                            # 获取新的列数
                            from openpyxl.utils import get_column_letter
                            new_col_count = len(header_row) + 1  # 多了一列
                            new_ref = f"A1:{get_column_letter(new_col_count)}{ws.max_row}"
                            table.ref = new_ref
                            print(f"    ✓ 更新 Table 范围: {new_ref}")

                        # 5. 保存迁移后的文件
                        wb.save(output_file)
                        print(f"  ✓ 列格式迁移完成！")

                        # 重新加载工作簿（确保后续操作使用更新后的结构）
                        wb = load_workbook(output_file)
                        ws = wb['Marks']

                    # 找到表格（如果存在）
                    table = None
                    if ws.tables:
                        table_name = list(ws.tables.keys())[0]
                        table = ws.tables[table_name]
                        print(f"  发现 Excel Table: {table_name}")

                    # 找到最后一行
                    last_row = ws.max_row

                    # 读取现有数据行数
                    existing_count = last_row - 1  # 减去表头

                    # 追加新数据（从最后一行的下一行开始）
                    for r_idx, row in enumerate(dataframe_to_rows(df_new, index=False, header=False), start=last_row + 1):
                        for c_idx, value in enumerate(row, start=1):
                            ws.cell(row=r_idx, column=c_idx, value=value)

                    # 如果有 Table，扩展其范围
                    if table:
                        # 计算新的表格范围
                        new_last_row = last_row + len(df_new)
                        # 获取列数
                        num_cols = len(df_new.columns)
                        # 更新表格范围
                        from openpyxl.utils import get_column_letter
                        new_ref = f"A1:{get_column_letter(num_cols)}{new_last_row}"
                        table.ref = new_ref
                        print(f"  已扩展 Table 范围到: {new_ref}")

                    # 保存
                    wb.save(output_file)
                    print(f"✓ 成功追加 {len(df_new)} 条新记录")
                    print(f"  总记录数: {existing_count + len(df_new)}")

                else:
                    # Sheet 不存在，创建新的
                    df_new.to_excel(output_file, sheet_name='Marks', index=False)
                    print(f"✓ 创建新表格，保存 {len(df_new)} 条记录")

            else:
                # 文件不存在，创建新文件
                df_new.to_excel(output_file, sheet_name='Marks', index=False)
                print(f"✓ 创建新文件，保存 {len(df_new)} 条记录")

            # 显示Excel列格式
            print("\nExcel 列格式:")
            print("  1. Batch Time - 批次开始时间")
            print("  2. Crawl Time - 商品爬取完成时间")
            print("  3. URL - 商品链接")
            print("  4. Price - 原价")
            print("  5. Promotion Price - 促销价")

        except Exception as e:
            print(f"✗ 保存失败: {str(e)}")
            # 保存到备份文件
            backup_file = f"Price_Marks_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            try:
                df_new.to_excel(backup_file, index=False)
                print(f"  已保存到备份文件: {backup_file}")
            except:
                pass

        print("\n" + "=" * 70)
        print("程序结束")
        print("=" * 70)

    except KeyboardInterrupt:
        print("\n\n程序被用户中断")
    except Exception as e:
        print(f"\n\n发生错误: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        print("\n正在关闭浏览器...")
        crawler.close()
        print("✓ 已关闭")


if __name__ == "__main__":
    main()
